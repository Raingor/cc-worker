"""Excel file reader — auto-detect structure, normalize to common format."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


@dataclass
class VersionSheet:
    """A version-comparison sheet (e.g. "04 23 Version")."""
    name: str
    total_skus: int | None = None
    total_qty: float | None = None
    total_amount: float | None = None
    headers: list[str] = field(default_factory=list)
    rows: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class MonthlyDemand:
    """Monthly demand sheet."""
    headers: list[str] = field(default_factory=list)
    rows: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class AnalysisFile:
    """Normalized result of parsing an analysis Excel file."""
    path: str
    filename: str
    version_sheets: list[VersionSheet] = field(default_factory=list)
    monthly_demand: MonthlyDemand | None = None
    file_type: str = "unknown"  # "sqrq_analysis" | "monthly_demand" | "unknown"


def _cell(val: Any) -> Any:
    """Clean a cell value: convert datetimes to ISO strings."""
    if isinstance(val, datetime):
        return val.isoformat()
    return val


def _is_date_heading(name: str) -> bool:
    """Detect if a sheet name looks like a version date: '04 23 Version'."""
    name = name.strip().lower()
    # Matches patterns like "04 23 Version", "04 17 Version " etc.
    parts = name.split()
    if len(parts) >= 2 and parts[-1] in ("version",):
        return True
    return False


def _is_summary_row(row: tuple) -> bool:
    """Check if a row looks like a summary (numeric tuple, not string headers)."""
    if not row or not row[0]:
        return False
    first = row[0]
    if isinstance(first, (int, float)):
        return True
    return False


def _find_header_row(ws, start_row: int = 1, max_check: int = 5) -> tuple[int, list[str]]:
    """Scan rows to find the real header row (looking for 'SKU' or similar)."""
    for r in range(start_row, min(start_row + max_check, ws.max_row + 1)):
        vals = [str(ws.cell(row=r, column=c).value or "") for c in range(1, ws.max_column + 1)]
        if any(kw in vals[0].lower() for kw in ("sku", "item", "物料", "料号")):
            return r, vals
    return start_row, []


def read_version_sheet(ws) -> VersionSheet:
    """Parse a single version comparison sheet."""
    sheet = VersionSheet(name=ws.title.strip())

    # Row 1: often a summary row
    r1 = tuple(ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1))
    if _is_summary_row(r1):
        sheet.total_skus = r1[0]
        sheet.total_qty = r1[1]
        sheet.total_amount = r1[2]

    # Find header row
    hr, headers = _find_header_row(ws, start_row=2)
    sheet.headers = [h for h in headers if h]

    # Data rows (after header row)
    for r in range(hr + 1, ws.max_row + 1):
        vals = tuple(ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1))
        if not any(v is not None for v in vals):
            continue  # skip empty rows
        row_dict = {}
        for i, h in enumerate(sheet.headers):
            if i < len(vals):
                row_dict[h] = _cell(vals[i])
        if row_dict.get("SKU"):
            sheet.rows.append(row_dict)

    return sheet


def read_monthly_demand(ws) -> MonthlyDemand:
    """Parse the Monthly Demand sheet."""
    md = MonthlyDemand()

    # Headers typically in row 3
    hr, headers = _find_header_row(ws, start_row=1)
    md.headers = [h for h in headers if h]

    for r in range(hr + 1, ws.max_row + 1):
        vals = tuple(ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1))
        if not any(v is not None for v in vals):
            continue
        row_dict = {}
        for i, h in enumerate(md.headers):
            if i < len(vals):
                row_dict[h] = _cell(vals[i])
        if row_dict.get("New SKU") or row_dict.get(md.headers[0] if md.headers else ""):
            md.rows.append(row_dict)

    return md


def read_file(filepath: str | Path) -> AnalysisFile:
    """Read an analysis Excel file and return structured data."""
    filepath = Path(filepath)
    result = AnalysisFile(
        path=str(filepath.resolve()),
        filename=filepath.name,
    )

    wb = openpyxl.load_workbook(filepath, data_only=True)

    for name in wb.sheetnames:
        ws = wb[name]
        name_clean = name.strip()

        if name_clean.lower() == "monthly demand":
            result.monthly_demand = read_monthly_demand(ws)
            continue

        if _is_date_heading(name_clean):
            result.version_sheets.append(read_version_sheet(ws))

    # Determine file type
    if result.version_sheets:
        result.file_type = "sqrq_analysis"
    elif result.monthly_demand:
        result.file_type = "monthly_demand"

    wb.close()
    return result
