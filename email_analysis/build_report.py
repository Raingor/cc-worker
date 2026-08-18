#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成 Mexicali CVR Analysis 报告: Excel(.xlsx) + PDF(.pdf)
复用 analyze.py 产出的图表 (charts/ 目录)
"""
import json
import numpy as np
import pandas as pd
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, Image as RLImage, PageBreak)

OUT = Path("/Users/mac-2312-r/workspace/wwwroot/CC/cc-worker/email_analysis")
CHART = OUT / "charts"
XLSX = OUT / "Mexicali_CVR_Analysis_Report.xlsx"
PDF = OUT / "Mexicali_CVR_Analysis_Report.pdf"

# ---------- 重新载入并计算 ----------
df = pd.read_excel(OUT / "1_锁盖分析 08 12 26.xlsx", sheet_name="Sheet1", header=1)
df.columns = ["Month", "v27639_25", "v27639_26",
              "v63251_25", "v63251_26", "v63252_25", "v63252_26"]
for c in df.columns:
    if c != "Month":
        df[c] = pd.to_numeric(df[c], errors="coerce")
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
df = df[df["Month"].isin(months)].reset_index(drop=True)
df["comb_25"] = df["v63251_25"] + df["v63252_25"]
df["comb_26"] = df["v63251_26"] + df["v63252_26"]

def stats(y25, y26, label):
    t25 = y25.sum(skipna=True); t26 = y26.sum(skipna=True)
    avg25 = y25.mean(skipna=True); avg26 = y26.mean(skipna=True)
    diff = t26 - t25
    pct = diff / t25 * 100 if t25 else float("nan")
    cv25 = y25.std(skipna=True) / avg25 * 100 if avg25 else float("nan")
    cv26 = y26.std(skipna=True) / avg26 * 100 if avg26 else float("nan")
    return dict(label=label, t25=t25, t26=t26, diff=diff, pct=pct,
                avg25=avg25, avg26=avg26, cv25=cv25, cv26=cv26,
                n25=int(y25.notna().sum()), n26=int(y26.notna().sum()))

s27639 = stats(df["v27639_25"], df["v27639_26"], "27639-003")
scomb   = stats(df["comb_25"], df["comb_26"], "63251-002 + 63252-002 (Combined)")
s63251  = stats(df["v63251_25"], df["v63251_26"], "63251-002")
s63252  = stats(df["v63252_25"], df["v63252_26"], "63252-002")

# ============================================================
# 1) Excel
# ============================================================
wb = Workbook()

# 样式
H = Font(bold=True, color="FFFFFF", size=11)
HF = PatternFill("solid", fgColor="305496")
TITLE = Font(bold=True, size=14, color="1F3864")
SUB = Font(bold=True, size=11, color="1F3864")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center")
NUM = "#,##0"
PCT = "+0.0%;-0.0%"

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H; cell.fill = HF; cell.alignment = CEN; cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# --- Sheet: Summary ---
ws = wb.active
ws.title = "Summary"
ws["A1"] = "Mexicali CVR Analysis — Summary Statistics"; ws["A1"].font = TITLE
ws["A2"] = "Source: 锁盖分析 08 12 26.xlsx  |  Comparison: 2025 vs 2026 order volume"
ws["A2"].font = Font(italic=True, color="808080")
hdr = ["Part Number", "2025 Total", "2026 Total", "Difference", "YoY Change",
       "2025 Avg/Month", "2026 Avg/Month", "CV 2025", "CV 2026", "Months(25/26)"]
r0 = 4
for j, h in enumerate(hdr, 1):
    ws.cell(row=r0, column=j, value=h)
style_header(ws, r0, len(hdr))
rows = [s27639, scomb, s63251, s63252]
for i, s in enumerate(rows, 1):
    r = r0 + i
    vals = [s["label"], s["t25"], s["t26"], s["diff"], s["pct"]/100,
            s["avg25"], s["avg26"], s["cv25"]/100, s["cv26"]/100,
            f'{s["n25"]}/{s["n26"]}']
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=j, value=v)
        cell.border = BORDER
        if j in (2,3,4,6,7): cell.number_format = NUM
        if j in (5,8,9): cell.number_format = PCT
        if j == 1: cell.font = Font(bold=True)
        if j in (2,3,4,5,6,7,8,9): cell.alignment = CEN
        if j == 5:
            cell.font = Font(bold=True, color=("C00000" if s["pct"] < 0 else "006100"))
autosize(ws, [34, 14, 14, 14, 12, 14, 14, 10, 10, 12])
ws.cell(row=r0+len(rows)+2, column=1,
        value="Note: 63251-002 Sept 2026 value is missing (NaN) in source; excluded from sums/averages.").font = Font(italic=True, color="C00000")

# --- Sheet: 27639 Monthly ---
ws = wb.create_sheet("27639_Monthly")
ws["A1"] = "27639-003 Monthly Detail (2025 vs 2026)"; ws["A1"].font = TITLE
hdr = ["Month", "2025", "2026", "Difference", "YoY Change"]
for j, h in enumerate(hdr, 1): ws.cell(row=3, column=j, value=h)
style_header(ws, 3, len(hdr))
for i, m in enumerate(months, 1):
    r = 3 + i
    a, b = df.loc[i-1, "v27639_25"], df.loc[i-1, "v27639_26"]
    d = (b - a) if (pd.notna(a) and pd.notna(b)) else None
    p = (d / a) if (d is not None and a) else None
    for j, v in enumerate([m, a, b, d, (p if p is not None else None)], 1):
        cell = ws.cell(row=r, column=j, value=v); cell.border = BORDER
        if j in (2,3,4): cell.number_format = NUM
        if j == 5 and p is not None:
            cell.number_format = PCT
            cell.font = Font(color=("C00000" if p < 0 else "006100"))
# totals row
tr = 3 + len(months) + 1
ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=tr, column=2, value=s27639["t25"]).number_format = NUM
ws.cell(row=tr, column=3, value=s27639["t26"]).number_format = NUM
ws.cell(row=tr, column=4, value=s27639["diff"]).number_format = NUM
ws.cell(row=tr, column=5, value=s27639["pct"]/100).number_format = PCT
for j in range(1,6): ws.cell(row=tr, column=j).border = BORDER
img = XLImage(str(CHART / "bar_27639.png")); img.width = 720; img.height = 327
ws.add_image(img, "G3")
autosize(ws, [10, 14, 14, 14, 12])

# --- Sheet: Combined Monthly ---
ws = wb.create_sheet("Combined_Monthly")
ws["A1"] = "63251-002 + 63252-002 Combined Monthly Detail"; ws["A1"].font = TITLE
hdr = ["Month", "2025 (sum)", "2026 (sum)", "Difference", "YoY Change"]
for j, h in enumerate(hdr, 1): ws.cell(row=3, column=j, value=h)
style_header(ws, 3, len(hdr))
for i, m in enumerate(months, 1):
    r = 3 + i
    a = df.loc[i-1, "comb_25"]
    b_raw = df.loc[i-1, "comb_26"]
    b = b_raw  # may include NaN from 63251 missing
    d = (b - a) if (pd.notna(a) and pd.notna(b)) else None
    p = (d / a) if (d is not None and a) else None
    for j, v in enumerate([m, a, b, d, (p if p is not None else None)], 1):
        cell = ws.cell(row=r, column=j, value=v); cell.border = BORDER
        if j in (2,3,4): cell.number_format = NUM
        if j == 5 and p is not None:
            cell.number_format = PCT
            cell.font = Font(color=("C00000" if p < 0 else "006100"))
    if pd.isna(b_raw):
        ws.cell(row=r, column=3).value = "N/A (63251 missing)"
tr = 3 + len(months) + 1
ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=tr, column=2, value=scomb["t25"]).number_format = NUM
ws.cell(row=tr, column=3, value=scomb["t26"]).number_format = NUM
ws.cell(row=tr, column=4, value=scomb["diff"]).number_format = NUM
ws.cell(row=tr, column=5, value=scomb["pct"]/100).number_format = PCT
for j in range(1,6): ws.cell(row=tr, column=j).border = BORDER
img = XLImage(str(CHART / "bar_combined.png")); img.width = 720; img.height = 327
ws.add_image(img, "G3")
autosize(ws, [10, 16, 20, 14, 12])

# --- Sheet: Fluctuation (3 SKU) ---
ws = wb.create_sheet("Fluctuation_3SKU")
ws["A1"] = "Monthly Fluctuation — 3 SKUs (2025 vs 2026)"; ws["A1"].font = TITLE
hdr = ["Month", "27639 25", "27639 26", "27639 Δ%",
       "63251 25", "63251 26", "63251 Δ%",
       "63252 25", "63252 26", "63252 Δ%"]
for j, h in enumerate(hdr, 1): ws.cell(row=3, column=j, value=h)
style_header(ws, 3, len(hdr))
for i, m in enumerate(months, 1):
    r = 3 + i
    row_vals = [m]
    for col in ["v27639", "v63251", "v63252"]:
        a = df.loc[i-1, f"{col}_25"]; b = df.loc[i-1, f"{col}_26"]
        d = (b - a) if (pd.notna(a) and pd.notna(b)) else None
        p = (d / a) if (d is not None and a) else None
        row_vals += [a, b, (p if p is not None else None)]
    for j, v in enumerate(row_vals, 1):
        cell = ws.cell(row=r, column=j, value=v); cell.border = BORDER
        if j % 3 == 2 or j % 3 == 0: cell.number_format = NUM
        if j % 3 == 0 and isinstance(v, (int, float)):
            cell.number_format = PCT
            cell.font = Font(color=("C00000" if v < 0 else "006100"))
# CV summary
sr = 3 + len(months) + 2
ws.cell(row=sr, column=1, value="Volatility (CV %)").font = SUB
cv_rows = [("27639-003", s27639), ("63251-002", s63251), ("63252-002", s63252),
           ("Combined", scomb)]
ws.cell(row=sr+1, column=1, value="SKU")
ws.cell(row=sr+1, column=2, value="CV 2025")
ws.cell(row=sr+1, column=3, value="CV 2026")
for j in (1,2,3): ws.cell(row=sr+1, column=j).font = H
for k, (lab, s) in enumerate(cv_rows, 1):
    ws.cell(row=sr+1+k, column=1, value=lab)
    ws.cell(row=sr+1+k, column=2, value=round(s["cv25"],1))
    ws.cell(row=sr+1+k, column=3, value=round(s["cv26"],1))
img = XLImage(str(CHART / "line_27639_fluct.png")); img.width = 720; img.height = 327
ws.add_image(img, "A" + str(sr + 7))
autosize(ws, [10, 12, 12, 10, 12, 12, 10, 12, 12, 10])

# --- Sheet: Raw Data ---
ws = wb.create_sheet("Raw_Data")
ws["A1"] = "Raw Data (as received)"; ws["A1"].font = TITLE
cols = ["Month","v27639_25","v27639_26","v63251_25","v63251_26","v63252_25","v63252_26"]
heads = ["Month","27639 2025","27639 2026","63251 2025","63251 2026","63252 2025","63252 2026"]
for j, h in enumerate(heads, 1): ws.cell(row=3, column=j, value=h)
style_header(ws, 3, len(heads))
for i, m in enumerate(months, 1):
    r = 3 + i
    for j, c in enumerate(cols, 1):
        v = df.loc[i-1, c]
        cell = ws.cell(row=r, column=j, value=(None if pd.isna(v) else v))
        cell.border = BORDER
        if j > 1: cell.number_format = NUM
ws.cell(row=3+len(months)+2, column=1,
        value="Note: 63251-002 2026 Sept = blank in source file.").font = Font(italic=True, color="C00000")
autosize(ws, [10, 14, 14, 14, 14, 14, 14])

wb.save(XLSX)
print("✅ Excel 已生成:", XLSX.name)

# ============================================================
# 2) PDF
# ============================================================
styles = getSampleStyleSheet()
H1 = ParagraphStyle("H1", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#1F3864"))
H2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13, textColor=colors.HexColor("#305496"))
BODY = ParagraphStyle("BODY", parent=styles["BodyText"], fontSize=10, leading=14)
SMALL = ParagraphStyle("SMALL", parent=styles["BodyText"], fontSize=8.5, textColor=colors.HexColor("#808080"))

doc = SimpleDocTemplate(str(PDF), pagesize=A4,
                        leftMargin=1.6*cm, rightMargin=1.6*cm,
                        topMargin=1.6*cm, bottomMargin=1.6*cm)
E = []

E.append(Paragraph("Mexicali CVR Analysis", H1))
E.append(Paragraph("Order Volume Comparison: 2025 vs 2026 &nbsp;|&nbsp; 3 SKUs", BODY))
E.append(Paragraph("Source: 锁盖分析 08 12 26.xlsx &nbsp;|&nbsp; Prepared for: Sylvia Tan", SMALL))
E.append(Spacer(1, 0.3*cm))

# Key findings
E.append(Paragraph("Key Findings", H2))
kf = [
    f"<b>27639-003 (analyzed separately):</b> total volume rose <b>+11.3%</b> "
    f"(13,641,000 → 15,180,000). Growth is led by Q1 (Jan +15.8%, Feb +56.3%, Apr +125.5%) "
    f"and a strong December (+80.3%). Monthly volatility (CV) improved from 79.6% to 59.4%.",
    f"<b>63251-002 + 63252-002 (combined):</b> total volume fell <b>-14.9%</b> "
    f"(3,344,400 → 2,847,000). Only Jan (+30%), Mar (+40%) and Nov (+233%, from a low base) grew; "
    f"Apr, May, Jul and Oct declined 25–49%.",
    f"<b>Data gap:</b> 63251-002 Sept 2026 is blank in the source file — excluded from sums and marked N/A.",
    f"<b>Pattern:</b> 27639 is a high-volume, seasonally-peaked item (January peak both years, "
    f"a blank March 2025 recovered in 2026); the two smaller SKUs show broad 2026 demand softening "
    f"outside Q1.",
]
for t in kf:
    E.append(Paragraph("• " + t, BODY))
E.append(Spacer(1, 0.2*cm))

# Summary table
E.append(Paragraph("Summary Statistics", H2))
data = [["Part Number", "2025", "2026", "Diff", "YoY", "CV25", "CV26"]]
for s in [s27639, scomb, s63251, s63252]:
    data.append([s["label"], f'{s["t25"]:,.0f}', f'{s["t26"]:,.0f}',
                 f'{s["diff"]:,.0f}', f'{s["pct"]:+.1f}%',
                 f'{s["cv25"]:.1f}%', f'{s["cv26"]:.1f}%'])
tbl = Table(data, colWidths=[6.0*cm, 2.6*cm, 2.6*cm, 2.6*cm, 1.8*cm, 1.6*cm, 1.6*cm])
tbl.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#305496")),
    ("TEXTCOLOR", (0,0), (-1,0), colors.white),
    ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
    ("FONTSIZE", (0,0), (-1,-1), 8.5),
    ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#BFBFBF")),
    ("ALIGN", (1,0), (-1,-1), "RIGHT"),
    ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#EEF1F7")]),
]))
E.append(tbl)
E.append(Spacer(1, 0.3*cm))

# Charts
E.append(Paragraph("Annual Total Comparison", H2))
E.append(RLImage(str(CHART / "bar_yearly_total.png"), width=15*cm, height=9.4*cm))
E.append(PageBreak())

E.append(Paragraph("27639-003 — Monthly Volume (2025 vs 2026)", H2))
E.append(RLImage(str(CHART / "bar_27639.png"), width=16*cm, height=7.3*cm))
E.append(Spacer(1, 0.2*cm))
E.append(Paragraph(
    "27639 monthly fluctuation (normalized to annual average): 2025 shows a sharp March gap (0 orders) "
    "and very high Jan/May/Jun peaks; 2026 is more evenly spread with a mid-year dip in May "
    "(2,142,000 → 480,000, -77.6%). CV dropped from 79.6% to 59.4%, i.e. 2026 is relatively more stable.",
    BODY))
E.append(RLImage(str(CHART / "line_27639_fluct.png"), width=16*cm, height=7.3*cm))
E.append(PageBreak())

E.append(Paragraph("63251-002 + 63252-002 — Combined Monthly Volume (2025 vs 2026)", H2))
E.append(RLImage(str(CHART / "bar_combined.png"), width=16*cm, height=7.3*cm))
E.append(Spacer(1, 0.2*cm))
E.append(Paragraph(
    "The two smaller SKUs contracted -14.9% in total. 2026 strength is concentrated in Q1 "
    "(Jan 489K→636K, Mar 205K→288K) and Nov (63K→210K); mid-year months are materially weaker "
    "(Apr -49%, May -44%, Jul -41%, Oct -28%). Sept 2026 for 63251 is missing (shown N/A); "
    "combined Sept 2026 = 63252 only (180,000). CV rose slightly from 48.2% to 54.2%.",
    BODY))
E.append(Spacer(1, 0.3*cm))

E.append(Paragraph("Possible Reasons (data-driven observations)", H2))
reasons = [
    "<b>Seasonality / build cycle:</b> 27639 peaks in January both years — consistent with an "
    "early-year demand or production build; the 2025 March blank suggests a one-off planning/data gap "
    "rather than a demand collapse (recovered in 2026).",
    "<b>Mid-year 2026 softness (27639):</b> May–Sep 2026 is below 2025 levels, implying a demand or "
    "scheduling shift in the 2026 mid-year — worth confirming against shipments/customer pulls.",
    "<b>Smaller SKU contraction:</b> broad 2026 declines across 63251/63252 outside Q1 point to "
    "overall demand softening, SKU substitution, or customer mix change for these parts.",
    "<b>Caveat:</b> root causes require business context (promotions, supply constraints, new "
    "programs). Figures above are derived solely from the supplied order-volume file.",
]
for t in reasons:
    E.append(Paragraph("• " + t, BODY))

doc.build(E)
print("✅ PDF 已生成:", PDF.name)
print("\n报告完成。")
